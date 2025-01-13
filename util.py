import os
import keyword
import traceback
import pandas, datetime
from sqlalchemy import inspect
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from db_config import Session, engine, Base, sql_reserved_keyword
from db_config.db_format import create_db_table, get_existing_table_class



def db_to_file( db_table, sheet_desc: str, sheet_title: str = "YCombinator Scraping Result", filename: str = 'ycombinator.com scraping result' ):
    ''' Method to save db into excel and csv file '''
    date = datetime.datetime.now().strftime('%d_%B_%Y')
    query = f"SELECT * FROM {db_table}"
    df = pandas.read_sql_query(query, engine)

    save_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
    os.makedirs(save_dir, exist_ok=True)

    excel_path = os.path.join(save_dir, f'{filename}_{date}.xlsx')
    csv_path = os.path.join(save_dir, f'{filename}_{date}.csv')

    with pandas.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=3, sheet_name='Sheet_1')
        worksheet = writer.sheets['Sheet_1']

        worksheet.merge_cells('A1:E1')
        worksheet.merge_cells('A2:E2')
        worksheet['A1'] = sheet_title
        worksheet['A2'] = sheet_desc

        worksheet['A1'].font = Font(name='Times New Roman', size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['A2'].alignment = Alignment(horizontal='left', vertical='center')

        print(f'Saved to file as: {filename}')

    df.to_csv(csv_path, index=False)
    print(f'File saved in {save_dir}')



def insert_to_db(scrape_result, table_class):
    """ This method used to insert data into database."""
    with Session() as session:
        table_instance = table_class(**scrape_result)
        session.add(table_instance)
        session.commit()


def check_data(data, table_class):
    """ Check if scraped data already in database """
    with Session() as session:
        exists =  session.query(table_class).filter_by(Name=data["Name"], Source_Url=data['Source_Url']).first()
        return exists is not None



def user_input_and_save_db_as_file(db_table= None):
    """ Function to input argument for saving database to file"""

    if db_table is not None:
        tablename = db_table
    else:
        while True:
            tablename = input('Input table name: ')
            if not check_table_exists(tablename):
                print("Table doesn't exist in database")
                continue
            break

    filename = input('Input filename (optional): ').strip() or None
    sheet_title = input('Input sheet title (optional): ').strip() or None
    sheet_desc = input('Input sheet description (optional) : ').strip() or (f'This file saved from table '
                                                                            f'{tablename} in ycombinator '
                                                                            f'database')

    args = {}
    if filename:
        args["filename"] = filename
    if sheet_title:
        args['sheet_title'] = sheet_title
    if sheet_desc:
        args['sheet_desc'] = sheet_desc
    args['db_table'] = tablename

    try:
        db_to_file(**args)
        return

    except Exception as e:
        print(f'Cannot save table to file: {e}')
        traceback.print_exc()
        return


def check_table_exists(tablename: str, engine=engine):
        inspector = inspect(engine)
        return tablename in inspector.get_table_names()


def user_input_scraping():

    while True:
        while True:
            url = input('Url: ')
            if not url.strip():
                print(" Url cannot be empty")
                continue
            if url_validation(url):
                break

        while True:
            count = input('Total Company (empty = all company): ')
            if count and not count.isdigit():
                print('Input must be a number')
                continue
            count = int(count) if count else None
            break

        while True:
            table_name = input('Input table name: ')
            valid = tablename_validator(table_name)

            if not valid:
                continue

            if check_table_exists(table_name):
                append = input('Table already exist, do you want to append data to the existing table (y/n): ').lower()
                if append not in ['y', 'n']:
                    print("Choose only y or n")
                    continue

                if append == 'y':
                    try:
                        table_class = get_existing_table_class(table_name)
                        print(f'Appending data to the existing table: {table_name}.')
                        return count, url, table_class, table_name
                    except ValueError as e:
                        print(f'Error on get existing table: {e}')
                        traceback.print_exc()
                        return
                else:
                    print('Please input the new table name!')
                    continue
            else:
                new_table_class = create_db_table(table_name)
                print(f'Table {table_name} has created ')
                return count, url, new_table_class, table_name



def url_validation(url) -> bool:
    try:
        result = urlparse(url)
        return all([result.scheme, result.netloc])
    except:
        return False


def tablename_validator(name) -> bool:

    if not name.strip():
        print('Table name cannot be empty')
        return False

    if not name.isidentifier():
        print('Table name must be a valid identifier')
        return False

    if name in keyword.kwlist or name.lower() in sql_reserved_keyword:
        print(f"Table name '{tablename}' is a reserved keyword and cannot be used")
        return False

    return True


def save_as_file_confirm():

    while True:
        save = input("Save result as file (y/n) : ").lower()
        if save and save in ['y', 'n']:
            if save == 'y':
               return True
            return False
        else:
            print('Choose only y or n')
            continue